# -*- coding: utf-8 -*-
# @Time    : 2019/9/2 10:40
# @Author  : project
# @File    : spider.py
# @Software: PyCharm
import re
import csv
import time
import hashlib
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


def run_time(func):
    def new_func(*args, **kwargs):
        start_time = datetime.now()
        print("程序开始时间：{}".format(start_time))
        func(*args, **kwargs)
        end_time = datetime.now()
        print("程序结束时间：{}".format(end_time))
        print("程序执行用时：{}s".format((end_time - start_time)))

    return new_func


class Spider:
    def __init__(self):
        options = Options()
        # 设置代理
        desired_capabilities = webdriver.DesiredCapabilities.INTERNETEXPLORER.copy()
        desired_capabilities['proxy'] = {
            "httpProxy": "https://127.0.0.1:1080",
            "proxyType": "MANUAL",  # 此项不可注释
        }
        # 使用无头模式
        # options.add_argument('headless')
        desired_capabilities = DesiredCapabilities.CHROME  # 修改页面加载策略
        desired_capabilities["pageLoadStrategy"] = "none"  # 注释这两行会导致最后输出结果的延迟，即等待页面加载完成再输出
        prefs = {"profile.managed_default_content_settings.images": 2}  # 1 加载图片 2不加载图片,加快访问速度
        options.add_experimental_option("prefs", prefs)
        # 此步骤很重要，设置为开发者模式，防止被各大网站识别出来使用了Selenium
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        # self.driver = webdriver.Chrome(chrome_options=options, desired_capabilities=desired_capabilities)
        self.driver = webdriver.Chrome(chrome_options=options)
        self.wait = WebDriverWait(self.driver, 5, 1)  # 设置隐式等待时间
        self.driver.maximize_window()  # 窗口最大化

        # 登录账号
        self.driver.get('https://twitter.com/login/error?redirect_after_login=%2Fashrafghani')
        input('账号登录')
        self.count = 1

    def loop_xlsx(self, header=True):
        """
        读取表格数据
        :return: ID URL
        """
        # 打开文件
        print('Open with openpyxl')
        workbook = openpyxl.load_workbook('config/用户id列表.xlsx')
        # 指定工作表
        sheet = 'Sheet1'
        sheet = workbook[workbook.sheetnames[sheet]] if isinstance(
            sheet, int) else workbook[sheet]

        # cell.value:单元格值 url:单元格超链接
        for row in sheet.iter_rows(min_row=1 + header):
            for c, cell in enumerate(row):
                url = cell.hyperlink.target if cell.hyperlink else None
                # print(' | ' if c else '', end='')
                if not cell.value:
                    return
                # print(f'{cell.value}{" (" + url + ")" if url else ""}', end='')
                yield cell.value, url

    def make_file_id(self, src):
        """
        生成哈希MD5码
        :param src: 字符串
        :return:
        """
        m1 = hashlib.md5()
        m1.update(src.encode('utf-8'))
        return m1.hexdigest()

    def friend_validation(self, make, filename):
        """
        效验是否获取过该条推文信息
        :return:
        """
        with open(f'config/{filename}', 'r') as f:
            flight = [i.replace('\n', '') for i in f.readlines()]
            if make in flight:
                return True
            with open(f'config/{filename}', 'a+') as f1:
                f1.write(make)
                f1.write('\n')
            return False

    def get_basic(self, acc_id, url):
        """
        获取基本信息: 推文数量、关注数、关注ID列表、粉丝数
        :return:
        """
        self.driver.get(url)
        while True:
            # try:
            time.sleep(3)
            # 发布文章数量
            Twitter_num = self.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="react-root"]/div/div/div/main/div/div/div/div[1]/div/div[1]/div/div/div/div/div/div[2]/div/div'))).text
            # print(Twitter_num)
            # 关注人数 关注列表页url
            # attention_num = self.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="react-root"]/div/div/div/main/div/div/div/div[1]/div/div[2]/div/div/div[1]/div/div[5]/div[1]/a/span[1]'))).text
            attention_num = '无'
            # print(attention_num)
            # attention_url = self.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="react-root"]/div/div/div/main/div/div/div/div[1]/div/div[2]/div/div/div[1]/div/div[5]/div[1]/a/@href')))
            attention_url = 'https://twitter.com/ashrafghani/following'
            # print(attention_url)
            # 粉丝
            # fan = self.wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="react-root"]/div/div/div/main/div/div/div/div[1]/div/div[2]/div/div/div[1]/div/div[5]/div[2]/a/span[1]/span'))).text
            fan = '104.1万'
            # print(fan)
            print(f'账号：{acc_id}, 发布推特数量：{Twitter_num}, 关注人数：{attention_num}, 粉丝：{fan}')
            #
            # # 获取关注ID列表
            # self.driver.get(attention_url)
            # time.sleep(3)
            # # 获取body的高度，滑到底部
            # scroll = "window.scrollTo(0,document.body.scrollHeight)"
            # self.driver.execute_script(scroll)
            #
            # attentions_id = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="react-root"]/div/div/div/main/div/div/div/div[1]/div/div[2]/section/div/div/div/div')))
            # attentions_id = [i.text for i in attentions_id if i.text]
            #
            # attention_id = []
            # for i in attentions_id:
            #     i = i.split('\n')
            #     attention_id.append(i[0])
            # print(f'关注列表：{attention_id}')
            content = [acc_id, Twitter_num, attention_num, fan, '无']
            return content
            # except:
            #     continue

    def get_teitter_content(self, url):
        """
        获取推特文本内容及相应获赞数量，评论数量，转发数量
        :return:
        """
        self.driver.get(url)
        count = 1
        while True:
            # 设置休眠时间 防止页面元素加载出错
            time.sleep(2)
            # 推文信息
            try:
                tweets = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="react-root"]/div/div/div/main/div/div/div/div[1]/div/div[2]/div/div/div[2]/section/div/div/div/div/div/article/div/div[2]/div[2]')))
                if not tweets:
                    continue
                tweets = [i.text.split('\n') for i in tweets if i.text]
                # print(tweets, len(tweets))  # 列表镶嵌列表

                # 效验是否获取过此条推文
                twee = [self.make_file_id(u) for i in tweets for u in i]
                # print(twee)
                # print(''.join(twee))
                make = self.make_file_id(''.join(twee))
                # print(make)
                if self.friend_validation(make, 'valida.txt'):
                    print('获取完成,退出，继续下一个!')
                    # 初始化计数
                    global COUNT
                    COUNT = 1
                    return
            except:
                # 滚动屏幕
                scroll = "window.scrollTo(0,document.body.scrollHeight)"
                self.driver.execute_script(scroll)
                count += 1
                continue
            yield tweets
            # 滚动屏幕
            scroll = "window.scrollTo(0,document.body.scrollHeight)"
            self.driver.execute_script(scroll)

    def processing(self, tweets):
        """
        数据处理
        tweets 列表镶嵌列表
        :return:
        """
        for tweet in tweets:
            content = tweet[4]
            release_time = tweet[3]
            comment = tweet[-3]
            forward = tweet[-2]
            like = tweet[-1]
            # 效验是否获取过此条推文
            make = self.make_file_id(content)
            if self.friend_validation(make, 'FriendValidation.txt'):
                print('数据重复,跳过!')
                continue
            yield content, release_time, comment, forward, like

    def save_data(self, data):
        """保存为csv"""
        with open("twitter.csv", "a+", encoding='utf-8', newline="") as f:
            k = csv.writer(f, delimiter=',')
            with open("twitter.csv", "r", encoding='utf-8', newline="") as f1:
                reader = csv.reader(f1)
                if not [row for row in reader]:
                    k.writerow(['ID', '推文数量', '关注人数', '关注列表', '粉丝数', '推文内容', '发布时间', '评论数', '转发数', '点赞数'])
                    k.writerows(data)
                else:
                    k.writerows(data)
            print(f'第：{self.count}条数据插入成功!')

    def quit(self):
        """
        关闭浏览器
        :return:
        """
        self.driver.quit()

    def run(self):
        for accunt in self.loop_xlsx():
            acc_id, url = accunt
            print(f'当前获取id：{acc_id} {url}')
            # 获取个人基本信息
            acc, Twitter_num, attention_num, fan, attention_id = self.get_basic(acc_id, url)
            # 推文数量
            Twitter_num = int(''.join(re.findall(r'\d', Twitter_num)))
            # 获取推文信息
            for tweets in self.get_teitter_content(url):
                # 提取信息
                for contents in self.processing(tweets):
                    content, release_time, comment, forward, like = contents
                    # 数据保存
                    data = [[acc, Twitter_num, attention_num, attention_id, fan, content, release_time, comment, forward, like]]
                    self.save_data(data)
                    self.count += 1
            # 清空效验文件
            with open('config/FriendValidation.txt', 'w') as f:
                f.seek(0)  # 光标移动至文件开头
                f.truncate()  # 清空文件
                print(f'ID:{acc_id} 效验文件清空成功')
            with open('config/valida.txt', 'w') as f:
                f.seek(0)  # 光标移动至文件开头
                f.truncate()  # 清空文件
                print(f'ID:{acc_id} 效验文件清空成功')
        self.quit()


def version():
    # 版本号
    print('*' * 20)
    print('脚本启动')
    print(f'selenium: {webdriver.__version__}')
    print(f'openpyxl: {openpyxl.__version__}')
    print('*' * 20)


if __name__ == '__main__':
    version()
    spider = Spider()
    spider.run()
