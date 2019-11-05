# -*- coding:utf-8 -*-
# 文件 ：计时器.py
# IED ：PyCharm
# 时间 ：2019/10/24 0024 9:03
# 版本 ：V1.0

import xlrd
import openpyxl


def loop_xls(workbook, sheet, header=True):
    sheet = workbook.sheet_by_index(0) if isinstance(
        sheet, int) else workbook.sheet_by_name(sheet)

    for r in range(int(header), sheet.nrows):
        for c, v in enumerate(sheet.row_values(r)):
            link = sheet.hyperlink_map.get((r, c))
            url = link.url_or_path if link else None

            print(' | ' if c else '', end='')
            print(f'{v}{" (" + url[:25] + ")" if url else ""}', end='')
        print()


def loop_xlsx(workbook, sheet, header=True):
    """
    cell.value:单元格值 url:单元格超链接
    """
    sheet = workbook[workbook.sheetnames[sheet]] if isinstance(
        sheet, int) else workbook[sheet]
    for row in sheet.iter_rows(min_row=1 + header):
        for c, cell in enumerate(row):
            url = cell.hyperlink.target if cell.hyperlink else None
            # print(' | ' if c else '', end='')
            if not cell.value:
                return
            print(f'{cell.value}{" (" + url + ")" if url else ""}', end='')
        print()


if __name__ == "__main__":
    # 版本号
    print(f'xlrd:     {xlrd.__version__}')
    print(f'openpyxl: {openpyxl.__version__}')
    print()

    files = [
        r'config/用户id列表.xlsx'
    ]

    for file in files:
        print(f'FILE: {file}')

        try:
            workbook = xlrd.open_workbook(file)
            print('Open with xlrd')
            loop_xls(workbook, 'url_table')
            print()
        except Exception:
            pass

        try:
            workbook = openpyxl.load_workbook(file)
            print('Open with openpyxl')
            loop_xlsx(workbook, 'Sheet1')
            print()
        except Exception as e:
            pass

# :: 注 ::
# xlrd 包虽能打开 .xlsx 文件，但不能正确读取超链接，并将整数读成浮点数
# openpyxl 包不能打开 .xls 文件，对于 .xlsx 文件中整数的处理也正确
